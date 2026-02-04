import openpyxl
import json

def extract_reference_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Budget']
    
    data = {
        "A.1 Story & Script": [],
        "B.1 Producers": [],
        "B.2 Directors": []
    }
    
    current_category = None
    
    # Simple state machine to capture specific sections
    # dependent on specific row text triggers found in previous analysis
    capture = False
    


    print("# Extracted Mockup Data (Comprehensive)\n")
    
    # We will scan for Category Headers which seem to be in Column A (index 0) looking like "A.1", "B.2", etc.
    # Or strict text matches in Column B/C if Column A is empty.
    
    # Corrected iter_rows call
    rows_iter = ws.iter_rows(min_row=1, values_only=True) 
    
    current_category = None
    data = {}
    
    # Regex for Category Codes like A.1, B.1, 10.1, etc.
    import re
    # Pattern: Start of string, 1-2 letters or digits, dot, 1-2 digits
    category_pattern = re.compile(r"^([A-Z0-9]{1,2}\.[0-9]{1,2})")

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Normalize row
        r = [str(cell).strip() if cell is not None else "" for cell in row]
        
        # Safety check for short rows
        if len(r) < 10:
            continue
            
        col0 = r[0] # Code often here
        col1 = r[1] # Sometimes here
        col2 = r[2] # Description often here
        
        # Check for Category Header
        # Usually Col0 has the code "A.1" and Col1 or Col2 has the Name
        match = category_pattern.match(col0)
        if match:
            # It's a header line!
            # Example: A.1, STORY & SCRIPT
            code = match.group(0)
            name = col1 if col1 else col2
            current_category = f"{code} {name}"
            data[current_category] = []
            continue
            
        # If we are in a category, let's look for valid line items
        if current_category:
            # Stop condition: Sub-total line or empty space that signifies end?
            # actually usually sub-total is part of the block, then a blank line.
            # identifying "Sub-total" is good to mark end of a block if we want, 
            # or just capture it.
            
            # Key data columns based on visual analysis:
            # Col 2: Desc
            # Col 3: Rate
            # Col 4: OT/Notes
            # Col 5: Unit 1
            desc = col2
            rate = r[3]
            amount1 = r[6]
            amount2 = r[8] 
            
            if "Sub-total" in col1 or "Sub-total" in col2:
                # We can skip or include subtotals. Let's include them for validity check.
                # data[current_category].append(r) 
                # actually, let's keep going until we hit a new category or end of sheet
                pass
            
            # If it has a Description OR a Rate/Amount, it's likely a line item
            # Avoid empty filler lines
            is_valid_line = (len(desc) > 1) or (len(rate) > 0 and rate != "Rate") or (len(amount1) > 0)
            
            if is_valid_line:
                # Filter out the header row repetition "Description... Rate..."
                if "Description" in desc and "Rate" in rate:
                    continue
                    
                data[current_category].append(r)

    # Output
    for cat, rows in data.items():
        if not rows: continue
        
        print(f"## {cat}")
        print("| Description | Rate | OT/Notes | Unit (Pre) | Amt (Pre) | Unit (Shoot) | Amt (Shoot) |")
        print("|---|---|---|---|---|---|---|")
        for r in rows:
            # Col 2: Desc
            # Col 3: R
            # Col 4: OT
            # Col 5: Unit1
            # Col 6: Amt1
            # Col 7: Unit2
            # Col 8: Amt2
            
            # Handle potential index errors if row is short (though we checked len < 10 above)
            d = r[2]
            rate = r[3]
            ot = r[4]
            u1 = r[5]
            a1 = r[6]
            u2 = r[7]
            a2 = r[8]
            
            print(f"| {d} | {rate} | {ot} | {u1} | {a1} | {u2} | {a2} |")
        print("\n")

if __name__ == "__main__":
    extract_reference_data("ShortKings_Standard Series Mockup Budget.xls.xlsx")
