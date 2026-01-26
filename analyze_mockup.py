"""
Analyze the mockup budget Excel file structure
"""
import openpyxl
import json

# Load workbook
wb = openpyxl.load_workbook('ShortKings_Standard Series Mockup Budget.xls.xlsx', data_only=True)

print("=" * 80)
print("MOCKUP BUDGET ANALYSIS")
print("=" * 80)

# List all sheets
print(f"\nTotal Sheets: {len(wb.sheetnames)}")
print("Sheets:", wb.sheetnames)

# Analyze Budget sheet
print("\n" + "=" * 80)
print("BUDGET SHEET STRUCTURE")
print("=" * 80)

ws = wb['Budget']
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns\n")

# Extract header row
print("HEADER ROW:")
headers = []
for col in range(1, min(20, ws.max_column + 1)):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value)
    print(f"  Col {col}: {cell_value}")

# Sample first 30 rows
print("\nFIRST 30 ROWS (showing first 10 columns):")
for row in range(1, min(31, ws.max_row + 1)):
    row_data = []
    for col in range(1, min(11, ws.max_column + 1)):
        val = ws.cell(row, col).value
        if val is not None:
            val_str = str(val)[:40]  # Truncate long values
            row_data.append(val_str)
        else:
            row_data.append("")
    print(f"Row {row:3d}: {row_data}")

# Analyze other key sheets
print("\n" + "=" * 80)
print("OTHER SHEET SAMPLES")
print("=" * 80)

for sheet_name in ['1.Dev', '2.Crew', '3.Cast', 'Summary']:
    if sheet_name in wb.sheetnames:
        print(f"\n--- {sheet_name} ---")
        ws = wb[sheet_name]
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print("First 10 rows (first 8 columns):")
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(9, ws.max_column + 1)):
                val = ws.cell(row, col).value
                if val is not None:
                    val_str = str(val)[:30]
                    row_data.append(val_str)
                else:
                    row_data.append("")
            print(f"  Row {row}: {row_data}")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
