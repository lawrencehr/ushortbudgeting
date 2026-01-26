
import openpyxl
import pandas as pd
import re

file_path = "ShortKings_Standard Series Mockup Budget.xls.xlsx"

def analyze_structure(file_path):
    print(f"Loading workbook: {file_path}...")
    try:
        # Load workbook to get formulas
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        print(f"Found {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")
        print("-" * 30)

        dependencies = {}
        
        # Regex to find sheet references like 'Sheet Name'!A1 or SheetName!A1
        # It's a bit complex to parse perfect Excel formulas, but a simple heuristic works for structure.
        # Look for something followed by '!'
        sheet_ref_pattern = re.compile(r"('?[^'!]+'?)\!")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            referenced_sheets = set()
            
            # Scan a reasonable range (e.g., first 200 rows, 20 cols) to find dependencies
            # We don't need to scan the whole million cells.
            for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        # It's a formula
                        matches = sheet_ref_pattern.findall(cell.value)
                        for match in matches:
                            # Clean up quotes
                            ref = match.strip("'")
                            if ref in wb.sheetnames and ref != sheet_name:
                                referenced_sheets.add(ref)
            
            dependencies[sheet_name] = list(referenced_sheets)
            
            if referenced_sheets:
                print(f"Sheet '{sheet_name}' pulls data from: {', '.join(referenced_sheets)}")
            else:
                print(f"Sheet '{sheet_name}' appears to have no external sheet links (in scanned range).")

        print("-" * 30)
        print("Detailed Sheet Inspection (First 5 rows):")
        
        # Use pandas for a pretty print of values
        # We assume the first row might be headers, but let's just print raw
        for sheet_name in wb.sheetnames:
            print(f"\n--- {sheet_name} ---")
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=5, engine='openpyxl')
            print(df.to_string(index=False, header=False))

    except Exception as e:
        print(f"Error analyzing workbook: {e}")

if __name__ == "__main__":
    analyze_structure(file_path)
